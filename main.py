from __future__ import annotations

from io import BytesIO
from pathlib import Path
import re

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import load_workbook


APP_TITLE = "Comparador de tendencias de formulación"
MAX_FILES = 12
MIN_FILES = 2
DEFAULT_INGREDIENT_METRIC = "inclusion_pct"
SPANISH_MONTHS = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").rstrip()


def to_float(token):
    if token is None:
        return None
    text = str(token).strip()
    if text in {"", ".", "+*.*", "*.*", "+.", "-", "--"}:
        return None
    text = text.replace(",", ".").replace("+", "")
    try:
        return float(text)
    except Exception:
        return None


def parse_period_from_line(line: str):
    match = re.search(
        r"(?<!\d)(Enero|Febrero|Marzo|Abril|Mayo|Junio|Julio|Agosto|Septiembre|Setiembre|Octubre|Noviembre|Diciembre)\s+(\d{2,4})",
        line,
        flags=re.IGNORECASE,
    )
    if not match:
        return None, None
    month_name = match.group(1).lower()
    year = int(match.group(2))
    if year < 100:
        year += 2000
    month_number = SPANISH_MONTHS.get(month_name)
    if not month_number:
        return None, None
    period_label = f"{match.group(1).capitalize()} {str(year)[-2:]}"
    period_date = pd.Timestamp(year=year, month=month_number, day=1)
    return period_label, period_date


def split_columns(line: str) -> list[str]:
    return [chunk.strip() for chunk in re.split(r"\s{2,}", line.strip()) if chunk.strip()]


def parse_product_descriptor(text: str):
    descriptor = re.sub(r"\s+", " ", text).strip()
    match = re.match(r"(?P<product_code>[A-Z0-9]+)\.(?P<product_name>.+)", descriptor)
    if match:
        return match.group("product_code").strip(), match.group("product_name").strip()
    parts = descriptor.split(None, 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return "", descriptor


def parse_ingredient_head(text: str):
    head = re.sub(r"\s+", " ", text).strip()
    parts = head.split(None, 1)
    local_code = parts[0].strip() if parts else ""
    rest = parts[1].strip() if len(parts) > 1 else ""
    match = re.match(r"(?P<ingredient_code>[A-Z0-9]+)\.(?P<ingredient_name>.+)", rest)
    if match:
        return local_code, match.group("ingredient_code").strip(), match.group("ingredient_name").strip()
    return local_code, "", rest


@st.cache_data(show_spinner=False)
def parse_uploaded_files(file_payloads: tuple[tuple[str, bytes], ...]):
    products_frames = []
    ingredients_frames = []
    nutrients_frames = []
    file_notes = []

    for file_name, file_bytes in file_payloads:
        try:
            products_df, ingredients_df, nutrients_df = parse_single_workbook(file_name, file_bytes)
            if products_df.empty:
                file_notes.append(f"{file_name}: no se detectaron bloques de formulación válidos.")
                continue
            products_frames.append(products_df)
            ingredients_frames.append(ingredients_df)
            nutrients_frames.append(nutrients_df)
        except Exception as exc:
            file_notes.append(f"{file_name}: error de lectura ({exc}).")

    if not products_frames:
        empty = pd.DataFrame()
        return {
            "products": empty,
            "ingredients": empty,
            "nutrients": empty,
            "notes": file_notes,
            "file_summary": empty,
            "period_order": [],
        }

    products = pd.concat(products_frames, ignore_index=True)
    ingredients = pd.concat(ingredients_frames, ignore_index=True) if ingredients_frames else pd.DataFrame()
    nutrients = pd.concat(nutrients_frames, ignore_index=True) if nutrients_frames else pd.DataFrame()

    file_summary = (
        products[["source_file", "period_label", "period_date"]]
        .drop_duplicates(subset=["source_file"])
        .sort_values(["period_date", "source_file"], kind="stable")
        .reset_index(drop=True)
    )
    duplicates = file_summary["period_label"].value_counts(dropna=False).to_dict()
    file_summary["period_display"] = file_summary.apply(
        lambda row: f"{row['period_label']} · {Path(row['source_file']).stem}"
        if duplicates.get(row["period_label"], 0) > 1
        else row["period_label"],
        axis=1,
    )
    display_map = dict(zip(file_summary["source_file"], file_summary["period_display"]))
    order_map = {label: pos for pos, label in enumerate(file_summary["period_display"].tolist())}

    for frame in (products, ingredients, nutrients):
        if frame.empty:
            continue
        frame["period_display"] = frame["source_file"].map(display_map)
        frame["period_order"] = frame["period_display"].map(order_map)
        frame.sort_values(["period_order"], inplace=True, kind="stable")

    return {
        "products": products,
        "ingredients": ingredients,
        "nutrients": nutrients,
        "notes": file_notes,
        "file_summary": file_summary,
        "period_order": file_summary["period_display"].tolist(),
    }


def parse_single_workbook(file_name: str, file_bytes: bytes):
    workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    lines = [clean_text(row[0]) for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True)]

    starts = [index for index, line in enumerate(lines) if "Specification:" in line]
    starts.append(len(lines))

    products = []
    ingredients = []
    nutrients = []

    for idx in range(len(starts) - 1):
        start = starts[idx]
        end = starts[idx + 1]
        block = lines[max(0, start - 10) : end]
        spec_line = lines[start]
        header_match = re.search(
            r"Specification:\s*(?P<spec>\S+)\s+(?P<descriptor>.+?)\s*:\s*Cost/tonne:\s*(?P<cost>[0-9.]+)",
            spec_line,
        )
        if not header_match:
            continue

        period_label = Path(file_name).stem
        period_date = pd.NaT
        for line in block[:15]:
            extracted_label, extracted_date = parse_period_from_line(line)
            if extracted_label:
                period_label = extracted_label
                period_date = extracted_date
                break

        total_tonnes = None
        for line in block[:10]:
            tonnes_match = re.search(r"Tonnes\s*:\s*([0-9.]+)", line)
            if tonnes_match:
                total_tonnes = to_float(tonnes_match.group(1))
                break

        spec_code = header_match.group("spec").strip()
        descriptor = re.sub(r"\s+", " ", header_match.group("descriptor")).strip()
        product_code, product_name = parse_product_descriptor(descriptor)
        product_key = f"{product_code} | {product_name}" if product_code else product_name
        product_display = f"{product_name} ({product_code})" if product_code else product_name
        cost_per_tonne = to_float(header_match.group("cost"))

        products.append(
            {
                "source_file": file_name,
                "period_label": period_label,
                "period_date": period_date,
                "spec_code": spec_code,
                "product_code": product_code,
                "product_name": product_name,
                "product_key": product_key,
                "product_display": product_display,
                "cost_per_tonne": cost_per_tonne,
                "total_tonnes": total_tonnes,
            }
        )

        raw_header_idx = next((i for i, line in enumerate(block) if "INCLUDED RAW MATERIALS" in line), None)
        analysis_header_idx = next((i for i, line in enumerate(block) if re.match(r"\s*ANALYSIS\b", line)), None)
        sensitivity_header_idx = next((i for i, line in enumerate(block) if "RAW MATERIAL SENSITIVITY" in line), None)

        if raw_header_idx is not None:
            raw_end = analysis_header_idx if analysis_header_idx is not None else len(block)
            for line in block[raw_header_idx + 2 : raw_end]:
                stripped = line.strip()
                if not stripped or stripped.startswith("-") or stripped.startswith("=") or "-------------" in stripped:
                    continue
                columns = split_columns(line)
                if len(columns) < 5:
                    continue
                local_code, ingredient_code, ingredient_name = parse_ingredient_head(columns[0])
                if not ingredient_name and not ingredient_code:
                    continue
                ingredient_key = f"{ingredient_code} | {ingredient_name}" if ingredient_code else ingredient_name
                ingredient_display = f"{ingredient_name} ({ingredient_code})" if ingredient_code else ingredient_name
                ingredients.append(
                    {
                        "source_file": file_name,
                        "period_label": period_label,
                        "period_date": period_date,
                        "spec_code": spec_code,
                        "product_code": product_code,
                        "product_name": product_name,
                        "product_key": product_key,
                        "product_display": product_display,
                        "ingredient_local_code": local_code,
                        "ingredient_code": ingredient_code,
                        "ingredient_name": ingredient_name,
                        "ingredient_key": ingredient_key,
                        "ingredient_display": ingredient_display,
                        "avg_cost": to_float(columns[1]) if len(columns) > 1 else None,
                        "inclusion_pct": to_float(columns[2]) if len(columns) > 2 else None,
                        "kilos": to_float(columns[3]) if len(columns) > 3 else None,
                        "tonnes": to_float(columns[4]) if len(columns) > 4 else None,
                        "limit_type": columns[5] if len(columns) > 5 else None,
                        "minimum": to_float(columns[6]) if len(columns) > 6 else None,
                        "maximum": to_float(columns[7]) if len(columns) > 7 else None,
                        "product_total_tonnes": total_tonnes,
                    }
                )

        if analysis_header_idx is not None:
            analysis_end = sensitivity_header_idx if sensitivity_header_idx is not None else len(block)
            for line in block[analysis_header_idx + 2 : analysis_end]:
                stripped = line.strip()
                if not stripped or stripped.startswith("-") or stripped.startswith("="):
                    continue
                columns = split_columns(line)
                if len(columns) < 2:
                    continue
                nutrient_name = columns[0]
                nutrients.append(
                    {
                        "source_file": file_name,
                        "period_label": period_label,
                        "period_date": period_date,
                        "spec_code": spec_code,
                        "product_code": product_code,
                        "product_name": product_name,
                        "product_key": product_key,
                        "product_display": product_display,
                        "nutrient_name": nutrient_name,
                        "level": to_float(columns[1]) if len(columns) > 1 else None,
                    }
                )

    return pd.DataFrame(products), pd.DataFrame(ingredients), pd.DataFrame(nutrients)


def weighted_average(values: pd.Series, weights: pd.Series):
    clean = pd.DataFrame({"value": values, "weight": weights}).dropna(subset=["value"])
    clean["weight"] = clean["weight"].fillna(0)
    positive_weight = clean["weight"].sum()
    if positive_weight > 0:
        return float((clean["value"] * clean["weight"]).sum() / positive_weight)
    if clean.empty:
        return None
    return float(clean["value"].mean())


def make_time_chart(dataframe: pd.DataFrame, value_col: str, series_col: str | None, title: str, y_axis_title: str):
    plot_df = dataframe.dropna(subset=[value_col]).copy()
    if plot_df.empty:
        st.info("No hay datos suficientes para la gráfica solicitada.")
        return

    sort_cols = ["period_order"]
    if series_col:
        sort_cols.append(series_col)
    plot_df = plot_df.sort_values(sort_cols, kind="stable")
    series_count = plot_df[series_col].nunique() if series_col else 1

    if series_col and series_count > 1:
        fig = px.line(
            plot_df,
            x="period_display",
            y=value_col,
            color=series_col,
            markers=True,
            title=title,
        )
    else:
        fig = px.bar(plot_df, x="period_display", y=value_col, title=title)

    fig.update_layout(
        xaxis_title="Periodo",
        yaxis_title=y_axis_title,
        legend_title="",
        margin=dict(l=20, r=20, t=60, b=20),
    )
    st.plotly_chart(fig, use_container_width=True)


def describe_change(series: pd.Series, unit: str = "", precision: int = 2):
    clean = series.dropna()
    if len(clean) < 2:
        return "sin datos suficientes para evaluar tendencia"
    start = float(clean.iloc[0])
    end = float(clean.iloc[-1])
    diff = end - start
    if abs(diff) < 1e-9:
        return f"estable ({end:.{precision}f}{unit})"
    direction = "al alza" if diff > 0 else "a la baja"
    if abs(start) > 1e-9:
        pct = diff / start * 100
        return f"{direction}: {start:.{precision}f}{unit} → {end:.{precision}f}{unit} ({pct:+.{precision}f}%)"
    return f"{direction}: {start:.{precision}f}{unit} → {end:.{precision}f}{unit}"


def build_product_report(
    products: pd.DataFrame,
    nutrients: pd.DataFrame,
    ingredients: pd.DataFrame,
    product_key: str,
    selected_nutrients: list[str],
    selected_ingredients: list[str],
    ingredient_metric: str,
):
    product_rows = products[products["product_key"] == product_key].copy().sort_values("period_order")
    if product_rows.empty:
        return "No hay datos del producto seleccionado."

    product_name = product_rows["product_display"].iloc[0]
    period_count = product_rows["period_display"].nunique()
    lines = [f"Informe para {product_name}.", f"Periodos analizados: {period_count}."]
    lines.append(
        "Precio del producto: "
        + describe_change(product_rows["cost_per_tonne"], unit=" €/t", precision=2)
        + "."
    )

    if selected_nutrients:
        nutrient_rows = nutrients[
            (nutrients["product_key"] == product_key) & (nutrients["nutrient_name"].isin(selected_nutrients))
        ].copy()
        if not nutrient_rows.empty:
            lines.append("Nutrientes seleccionados:")
            for nutrient_name, group in nutrient_rows.groupby("nutrient_name"):
                group = group.sort_values("period_order")
                latest_value = group["level"].dropna().iloc[-1] if not group["level"].dropna().empty else None
                latest_text = f" Último valor: {latest_value:.3f}." if latest_value is not None else ""
                lines.append(f"- {nutrient_name}: {describe_change(group['level'], precision=3)}.{latest_text}")

    if selected_ingredients:
        metric_label = {
            "inclusion_pct": "% inclusión",
            "kilos": "kg por lote",
            "tonnes": "t por lote",
        }[ingredient_metric]
        ingredient_rows = ingredients[
            (ingredients["product_key"] == product_key) & (ingredients["ingredient_key"].isin(selected_ingredients))
        ].copy()
        if not ingredient_rows.empty:
            lines.append(f"Ingredientes seleccionados ({metric_label}):")
            for ingredient_name, group in ingredient_rows.groupby("ingredient_display"):
                group = group.sort_values("period_order")
                lines.append(f"- {ingredient_name}: {describe_change(group[ingredient_metric], precision=3)}.")

    return "\n".join(lines)


def build_ingredient_report(
    ingredient_display: str,
    price_series: pd.DataFrame,
    consumption_series: pd.DataFrame,
):
    lines = [f"Informe para {ingredient_display}."]
    lines.append(
        "Precio estimado del ingrediente: "
        + describe_change(price_series["value"], unit=" €/t", precision=2)
        + "."
    )
    lines.append(
        "Consumo relativo estimado en los archivos cargados: "
        + describe_change(consumption_series["ingredient_tonnes"], unit=" t", precision=3)
        + "."
    )
    lines.append(
        "La estimación de consumo es relativa: suma las toneladas presentes en las formulaciones cargadas y no sustituye a un dato real de fabricación o compras."
    )
    return "\n".join(lines)


def build_export_workbook(report_text: str, sheets: dict[str, pd.DataFrame]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        report_lines = pd.DataFrame({"informe": report_text.splitlines()})
        report_lines.to_excel(writer, sheet_name="Resumen", index=False)
        for sheet_name, dataframe in sheets.items():
            export_df = dataframe.copy()
            for column in export_df.columns:
                if pd.api.types.is_datetime64_any_dtype(export_df[column]):
                    export_df[column] = export_df[column].dt.strftime("%Y-%m-%d")
            safe_name = re.sub(r"[^A-Za-z0-9_]", "_", sheet_name)[:31] or "Hoja"
            export_df.to_excel(writer, sheet_name=safe_name, index=False)
    return output.getvalue()


def read_readme() -> str:
    readme_path = Path(__file__).with_name("README.md")
    if readme_path.exists():
        return readme_path.read_text(encoding="utf-8")
    return "README no disponible."


def get_sorted_unique_options(dataframe: pd.DataFrame, key_col: str, label_col: str):
    options = dataframe[[key_col, label_col]].drop_duplicates().sort_values(label_col, kind="stable")
    return options[key_col].tolist(), dict(zip(options[key_col], options[label_col]))


def build_nutrient_wide_table(nutrient_series: pd.DataFrame, selected_nutrients: list[str]) -> pd.DataFrame:
    if nutrient_series.empty:
        return pd.DataFrame()

    index_cols = ["period_display", "period_order", "period_date", "source_file"]
    pivot = (
        nutrient_series[index_cols + ["nutrient_name", "level"]]
        .pivot_table(index=index_cols, columns="nutrient_name", values="level", aggfunc="first")
        .reset_index()
        .sort_values(["period_order", "period_date", "source_file"], kind="stable")
    )

    ordered_nutrient_cols = [name for name in selected_nutrients if name in pivot.columns]
    return pivot[["period_display", *ordered_nutrient_cols, "source_file"]]


def clear_parse_cache():
    if hasattr(parse_uploaded_files, "clear"):
        parse_uploaded_files.clear()


def initialize_session_state():
    defaults = {
        "uploader_key": 0,
        "selected_product": "",
        "selected_nutrients_ui": [],
        "selected_ingredients_ui": [],
        "ingredient_metric_ui": DEFAULT_INGREDIENT_METRIC,
        "last_product_for_widgets": None,
        "nutrient_memory_by_product": {},
        "ingredient_memory_by_product": {},
        "metric_memory_by_product": {},
        "global_ingredient_selection": [],
        "global_ingredient_metric": DEFAULT_INGREDIENT_METRIC,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def reset_analysis_state():
    next_uploader_key = st.session_state.get("uploader_key", 0) + 1
    keys_to_reset = [
        "selected_product",
        "selected_nutrients_ui",
        "selected_ingredients_ui",
        "ingredient_metric_ui",
        "last_product_for_widgets",
        "nutrient_memory_by_product",
        "ingredient_memory_by_product",
        "metric_memory_by_product",
        "global_ingredient_selection",
        "global_ingredient_metric",
    ]
    for key in keys_to_reset:
        st.session_state.pop(key, None)
    st.session_state["uploader_key"] = next_uploader_key
    initialize_session_state()


def sync_selection_state(current_product: str, nutrient_options: list[str], ingredient_option_keys: list[str]):
    last_product = st.session_state.get("last_product_for_widgets")

    if last_product != current_product:
        if current_product:
            st.session_state["selected_nutrients_ui"] = [
                nutrient
                for nutrient in st.session_state["nutrient_memory_by_product"].get(current_product, [])
                if nutrient in nutrient_options
            ]
            st.session_state["selected_ingredients_ui"] = [
                ingredient
                for ingredient in st.session_state["ingredient_memory_by_product"].get(current_product, [])
                if ingredient in ingredient_option_keys
            ]
            st.session_state["ingredient_metric_ui"] = st.session_state["metric_memory_by_product"].get(
                current_product,
                DEFAULT_INGREDIENT_METRIC,
            )
        else:
            st.session_state["selected_nutrients_ui"] = []
            st.session_state["selected_ingredients_ui"] = [
                ingredient
                for ingredient in st.session_state.get("global_ingredient_selection", [])
                if ingredient in ingredient_option_keys
            ]
            st.session_state["ingredient_metric_ui"] = st.session_state.get(
                "global_ingredient_metric",
                DEFAULT_INGREDIENT_METRIC,
            )

        st.session_state["last_product_for_widgets"] = current_product
    else:
        st.session_state["selected_nutrients_ui"] = [
            nutrient for nutrient in st.session_state.get("selected_nutrients_ui", []) if nutrient in nutrient_options
        ]
        st.session_state["selected_ingredients_ui"] = [
            ingredient
            for ingredient in st.session_state.get("selected_ingredients_ui", [])
            if ingredient in ingredient_option_keys
        ]


def persist_current_selection(selected_product: str, selected_nutrients: list[str], selected_ingredients: list[str], ingredient_metric: str):
    if selected_product:
        st.session_state["nutrient_memory_by_product"][selected_product] = list(selected_nutrients)
        st.session_state["ingredient_memory_by_product"][selected_product] = list(selected_ingredients)
        st.session_state["metric_memory_by_product"][selected_product] = ingredient_metric
    else:
        st.session_state["global_ingredient_selection"] = list(selected_ingredients)
        st.session_state["global_ingredient_metric"] = ingredient_metric


def main():
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    initialize_session_state()

    st.title(APP_TITLE)
    st.caption(
        "Carga entre 2 y 12 archivos Excel del mismo formato para analizar precios, nutrientes e ingredientes por periodo."
    )

    with st.sidebar:
        st.header("Carga de archivos")
        action_cols = st.columns(2)
        if action_cols[0].button("Refrescar", use_container_width=True):
            clear_parse_cache()
            st.rerun()
        if action_cols[1].button("Nuevo análisis", use_container_width=True):
            reset_analysis_state()
            clear_parse_cache()
            st.rerun()

        uploaded_files = st.file_uploader(
            "Selecciona archivos .xlsx",
            type=["xlsx"],
            accept_multiple_files=True,
            key=f"file_uploader_{st.session_state['uploader_key']}",
            help="Formato esperado: libro con una hoja y bloques de formulación como el archivo de ejemplo.",
        )
        st.markdown(f"Mínimo: **{MIN_FILES}** archivos · Máximo: **{MAX_FILES}** archivos")

    tabs = st.tabs(["Análisis", "Datos", "README"])

    with tabs[2]:
        st.markdown(read_readme())

    if not uploaded_files:
        with tabs[0]:
            st.info("Carga archivos para empezar.")
        with tabs[1]:
            st.info("Todavía no hay datos cargados.")
        return

    if len(uploaded_files) > MAX_FILES:
        with tabs[0]:
            st.error(f"Se han cargado {len(uploaded_files)} archivos. El máximo admitido es {MAX_FILES}.")
        with tabs[1]:
            st.info("Reduce el número de archivos para continuar.")
        return

    payloads = tuple((uploaded_file.name, uploaded_file.getvalue()) for uploaded_file in uploaded_files)
    parsed = parse_uploaded_files(payloads)
    products = parsed["products"].copy()
    ingredients = parsed["ingredients"].copy()
    nutrients = parsed["nutrients"].copy()
    file_summary = parsed["file_summary"].copy()

    with tabs[0]:
        if products.empty:
            st.error("No se han podido extraer datos útiles de los archivos cargados.")
            return

        metric_cols = st.columns(4)
        metric_cols[0].metric("Archivos válidos", int(file_summary.shape[0]))
        metric_cols[1].metric("Productos detectados", int(products["product_key"].nunique()))
        metric_cols[2].metric(
            "Ingredientes detectados",
            int(ingredients["ingredient_key"].nunique()) if not ingredients.empty else 0,
        )
        metric_cols[3].metric(
            "Nutrientes detectados",
            int(nutrients["nutrient_name"].nunique()) if not nutrients.empty else 0,
        )

        if parsed["notes"]:
            for note in parsed["notes"]:
                st.warning(note)

        if len(file_summary) < MIN_FILES:
            st.warning(
                f"Para comparar tendencias deben cargarse al menos {MIN_FILES} archivos. Ahora mismo hay {len(file_summary)} válido(s)."
            )

        product_keys, product_labels = get_sorted_unique_options(products, "product_key", "product_display")
        if not ingredients.empty:
            ingredient_keys, ingredient_labels = get_sorted_unique_options(ingredients, "ingredient_key", "ingredient_display")
        else:
            ingredient_keys, ingredient_labels = [], {}

        product_options = [""] + product_keys
        if st.session_state.get("selected_product", "") not in product_options:
            st.session_state["selected_product"] = ""

        current_product = st.session_state.get("selected_product", "")
        nutrient_scope = nutrients[nutrients["product_key"] == current_product] if current_product else nutrients
        nutrient_options = sorted(nutrient_scope["nutrient_name"].dropna().unique().tolist()) if not nutrient_scope.empty else []

        ingredient_scope = ingredients[ingredients["product_key"] == current_product] if current_product else ingredients
        ingredient_options_df = (
            ingredient_scope[["ingredient_key", "ingredient_display"]]
            .drop_duplicates()
            .sort_values("ingredient_display", kind="stable")
            if not ingredient_scope.empty
            else pd.DataFrame(columns=["ingredient_key", "ingredient_display"])
        )
        ingredient_option_keys = ingredient_options_df["ingredient_key"].tolist()
        ingredient_option_labels = dict(
            zip(ingredient_options_df["ingredient_key"], ingredient_options_df["ingredient_display"])
        )

        sync_selection_state(current_product, nutrient_options, ingredient_option_keys)

        selector_cols = st.columns([1.3, 1, 1])
        selected_product = selector_cols[0].selectbox(
            "Producto",
            options=product_options,
            key="selected_product",
            format_func=lambda key: "— Ninguno —" if key == "" else product_labels.get(key, key),
        )

        selected_nutrients = selector_cols[1].multiselect(
            "Nutrientes",
            options=nutrient_options,
            key="selected_nutrients_ui",
        )

        selected_ingredients = selector_cols[2].multiselect(
            "Ingredientes",
            options=ingredient_option_keys,
            key="selected_ingredients_ui",
            format_func=lambda key: ingredient_option_labels.get(key, key),
        )

        ingredient_metric = st.radio(
            "Métrica para ingredientes en vista de producto",
            options=["inclusion_pct", "kilos", "tonnes"],
            key="ingredient_metric_ui",
            format_func=lambda value: {
                "inclusion_pct": "% inclusión",
                "kilos": "kg por lote",
                "tonnes": "t por lote",
            }[value],
            horizontal=True,
        )

        persist_current_selection(selected_product, selected_nutrients, selected_ingredients, ingredient_metric)

        report_text = "Selecciona un producto o un único ingrediente para generar el informe."
        export_sheets: dict[str, pd.DataFrame] = {
            "Productos_cargados": products,
            "Ingredientes_cargados": ingredients,
            "Nutrientes_cargados": nutrients,
            "Archivos": file_summary,
        }

        if selected_product:
            st.subheader("Vista por producto")
            product_series = products[products["product_key"] == selected_product].copy().sort_values(
                ["period_order", "period_date", "source_file"],
                kind="stable",
            )
            make_time_chart(
                product_series,
                value_col="cost_per_tonne",
                series_col=None,
                title="Evolución temporal del precio del producto",
                y_axis_title="€/t",
            )
            st.dataframe(
                product_series[["period_display", "product_display", "cost_per_tonne", "total_tonnes", "source_file"]]
                .rename(
                    columns={
                        "period_display": "Periodo",
                        "product_display": "Producto",
                        "cost_per_tonne": "Precio €/t",
                        "total_tonnes": "Toneladas lote",
                        "source_file": "Archivo",
                    }
                ),
                use_container_width=True,
                hide_index=True,
            )
            export_sheets["Precio_producto"] = product_series

            if selected_nutrients:
                st.subheader("Nutrientes seleccionados")
                nutrient_series = nutrients[
                    (nutrients["product_key"] == selected_product) & (nutrients["nutrient_name"].isin(selected_nutrients))
                ].copy()
                nutrient_series.sort_values(["period_order", "period_date", "nutrient_name"], inplace=True)
                make_time_chart(
                    nutrient_series,
                    value_col="level",
                    series_col="nutrient_name",
                    title="Evolución temporal de nutrientes",
                    y_axis_title="Nivel",
                )
                nutrient_wide = build_nutrient_wide_table(nutrient_series, selected_nutrients)
                st.dataframe(
                    nutrient_wide.rename(
                        columns={
                            "period_display": "Periodo",
                            "source_file": "Archivo",
                        }
                    ),
                    use_container_width=True,
                    hide_index=True,
                )
                export_sheets["Nutrientes_seleccionados"] = nutrient_wide

            if selected_ingredients:
                st.subheader("Ingredientes seleccionados")
                ingredient_series = ingredients[
                    (ingredients["product_key"] == selected_product)
                    & (ingredients["ingredient_key"].isin(selected_ingredients))
                ].copy()
                ingredient_series.sort_values(["period_order", "ingredient_display"], inplace=True)
                make_time_chart(
                    ingredient_series,
                    value_col=ingredient_metric,
                    series_col="ingredient_display",
                    title="Evolución temporal de ingredientes en el producto",
                    y_axis_title={
                        "inclusion_pct": "% inclusión",
                        "kilos": "kg por lote",
                        "tonnes": "t por lote",
                    }[ingredient_metric],
                )
                st.dataframe(
                    ingredient_series[
                        [
                            "period_display",
                            "ingredient_display",
                            "avg_cost",
                            "inclusion_pct",
                            "kilos",
                            "tonnes",
                            "source_file",
                        ]
                    ].rename(
                        columns={
                            "period_display": "Periodo",
                            "ingredient_display": "Ingrediente",
                            "avg_cost": "Precio €/t",
                            "inclusion_pct": "% inclusión",
                            "kilos": "Kg",
                            "tonnes": "Toneladas",
                            "source_file": "Archivo",
                        }
                    ),
                    use_container_width=True,
                    hide_index=True,
                )
                export_sheets["Ingredientes_seleccionados"] = ingredient_series

            report_text = build_product_report(
                products=products,
                nutrients=nutrients,
                ingredients=ingredients,
                product_key=selected_product,
                selected_nutrients=selected_nutrients,
                selected_ingredients=selected_ingredients,
                ingredient_metric=ingredient_metric,
            )

        elif len(selected_ingredients) == 1:
            st.subheader("Vista por ingrediente")
            ingredient_key = selected_ingredients[0]
            ingredient_rows = ingredients[ingredients["ingredient_key"] == ingredient_key].copy()
            ingredient_display = ingredient_rows["ingredient_display"].iloc[0]

            price_series = (
                ingredient_rows.groupby(["period_display", "period_order", "period_date"], dropna=False)
                .apply(lambda group: weighted_average(group["avg_cost"], group["tonnes"]))
                .reset_index(name="value")
                .sort_values("period_order", kind="stable")
            )
            make_time_chart(
                price_series,
                value_col="value",
                series_col=None,
                title="Evolución del precio del ingrediente",
                y_axis_title="€/t",
            )
            st.dataframe(
                price_series.rename(
                    columns={
                        "period_display": "Periodo",
                        "value": "Precio estimado €/t",
                    }
                )[["Periodo", "Precio estimado €/t"]],
                use_container_width=True,
                hide_index=True,
            )

            denominator = (
                products.groupby(["period_display", "period_order", "period_date"], dropna=False)["total_tonnes"]
                .sum(min_count=1)
                .reset_index(name="loaded_total_tonnes")
            )
            consumption_series = (
                ingredient_rows.groupby(["period_display", "period_order", "period_date"], dropna=False)["tonnes"]
                .sum(min_count=1)
                .reset_index(name="ingredient_tonnes")
                .merge(denominator, on=["period_display", "period_order", "period_date"], how="left")
            )
            consumption_series["share_pct"] = (
                consumption_series["ingredient_tonnes"] / consumption_series["loaded_total_tonnes"] * 100
            )
            make_time_chart(
                consumption_series,
                value_col="ingredient_tonnes",
                series_col=None,
                title="Estimación relativa del consumo del ingrediente",
                y_axis_title="Toneladas presentes en formulaciones cargadas",
            )
            st.dataframe(
                consumption_series.rename(
                    columns={
                        "period_display": "Periodo",
                        "ingredient_tonnes": "Toneladas ingrediente",
                        "loaded_total_tonnes": "Toneladas totales cargadas",
                        "share_pct": "% sobre formulaciones cargadas",
                    }
                )[
                    [
                        "Periodo",
                        "Toneladas ingrediente",
                        "Toneladas totales cargadas",
                        "% sobre formulaciones cargadas",
                    ]
                ],
                use_container_width=True,
                hide_index=True,
            )

            report_text = build_ingredient_report(
                ingredient_display=ingredient_display,
                price_series=price_series,
                consumption_series=consumption_series,
            )
            export_sheets["Precio_ingrediente"] = price_series
            export_sheets["Consumo_ingrediente"] = consumption_series

        else:
            st.info(
                "Para la vista analítica, selecciona un producto o deja el producto vacío y marca un único ingrediente."
            )

        st.subheader("Informe de texto")
        st.text_area("Resumen del análisis", report_text, height=260)

        excel_bytes = build_export_workbook(report_text, export_sheets)
        download_cols = st.columns(2)
        download_cols[0].download_button(
            "Descargar análisis en Excel",
            data=excel_bytes,
            file_name="analisis_formulacion.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        download_cols[1].download_button(
            "Descargar informe de texto",
            data=report_text.encode("utf-8"),
            file_name="informe_analisis.txt",
            mime="text/plain",
        )

    with tabs[1]:
        st.subheader("Resumen de archivos")
        st.dataframe(file_summary, use_container_width=True, hide_index=True)

        st.subheader("Productos")
        st.dataframe(products, use_container_width=True, hide_index=True)

        st.subheader("Ingredientes")
        st.dataframe(ingredients, use_container_width=True, hide_index=True)

        st.subheader("Nutrientes")
        st.dataframe(nutrients, use_container_width=True, hide_index=True)


if __name__ == "__main__":
    main()
